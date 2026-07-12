# -*- coding: utf-8 -*-
"""
错标追查 + 工单查询 + 库位搜索 - 路由模块
包含：/blame_query, /api/blame_query, /api/ticket_query,
      /api/ticket_detail, /api/sku_location_search
"""
import os
import re
import glob
from pathlib import Path
from flask import Blueprint, render_template, jsonify, request
from datetime import datetime, timedelta
from token_manager import get_access_token
from spiders.base import request_with_retry, get_shared_session

bp = Blueprint('blame', __name__)

# ============================================================
# SKU 级别拣货时间缓存（从下载目录的导出xlsx文件解析）
# ============================================================
_sku_pick_time_cache = {}  # key: (子任务号, sku编码) -> 商品拣货时间字符串
_cache_loaded_files = set()  # 已加载过的文件路径集合
_DOWNLOADS_DIR = str(Path.home() / "Downloads")
_CACHE_DAYS = 4  # 扫描近4天的文件


def _load_zonepick_export_cache():
    """扫描下载目录中近4天的储位拣货子任务导出详情xlsx文件，解析并缓存SKU拣货时间"""
    global _sku_pick_time_cache, _cache_loaded_files

    try:
        import openpyxl
    except ImportError:
        print("[SKU拣货时间缓存] openpyxl未安装，无法解析xlsx文件")
        return

    # 查找匹配的文件
    pattern = os.path.join(_DOWNLOADS_DIR, "储位拣货子任务导出详情*xlsx")
    all_files = glob.glob(pattern)

    if not all_files:
        return

    # 筛选近4天内修改的文件
    cutoff_time = datetime.now() - timedelta(days=_CACHE_DAYS)
    recent_files = []
    for f in all_files:
        try:
            mtime = datetime.fromtimestamp(os.path.getmtime(f))
            if mtime >= cutoff_time:
                recent_files.append(f)
        except OSError:
            continue

    if not recent_files:
        return

    new_count = 0
    for filepath in recent_files:
        if filepath in _cache_loaded_files:
            continue  # 已加载过，跳过

        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            ws = wb.active

            # 读取表头，找到关键列的索引
            headers = [cell.value for cell in ws[1]]
            col_task_no = None  # 拣货子任务号
            col_sku_code = None  # SKU编码
            col_pick_time = None  # 商品拣货时间

            for idx, h in enumerate(headers):
                if h == "拣货子任务号":
                    col_task_no = idx
                elif h == "SKU编码":
                    col_sku_code = idx
                elif h == "商品拣货时间":
                    col_pick_time = idx

            if col_task_no is None or col_sku_code is None or col_pick_time is None:
                wb.close()
                continue

            # 逐行解析
            for row in ws.iter_rows(min_row=2, values_only=True):
                task_no = row[col_task_no] if col_task_no < len(row) else None
                sku_code = row[col_sku_code] if col_sku_code < len(row) else None
                pick_time = row[col_pick_time] if col_pick_time < len(row) else None

                if task_no and sku_code and pick_time:
                    # SKU编码统一为字符串（去掉小数点）
                    sku_str = str(int(float(sku_code))) if isinstance(sku_code, (int, float)) else str(sku_code).strip()
                    # 时间格式化
                    if isinstance(pick_time, datetime):
                        time_str = pick_time.strftime("%Y-%m-%d %H:%M:%S")
                    else:
                        time_str = str(pick_time).strip()

                    cache_key = (str(task_no).strip(), sku_str)
                    _sku_pick_time_cache[cache_key] = time_str
                    new_count += 1

            wb.close()
            _cache_loaded_files.add(filepath)
        except Exception as e:
            print(f"[SKU拣货时间缓存] 解析文件失败 {filepath}: {e}")
            continue

    if new_count > 0:
        print(f"[SKU拣货时间缓存] 已加载 {len(_cache_loaded_files)} 个文件，共 {len(_sku_pick_time_cache)} 条缓存记录")


def get_sku_pick_time(zone_pick_bill_no, sku_code):
    """从缓存中查询指定子任务号+SKU的商品拣货时间
    
    Args:
        zone_pick_bill_no: 拣货子任务号（如 DBJ4282026060400004-1）
        sku_code: SKU编码
    
    Returns:
        (时间字符串, 是否找到) 元组
    """
    # 每次查询时刷新缓存（检查是否有新文件）
    _load_zonepick_export_cache()

    if not _sku_pick_time_cache:
        return None, False

    sku_str = str(int(float(sku_code))) if sku_code else ""
    task_no_str = str(zone_pick_bill_no).strip() if zone_pick_bill_no else ""

    # 精确匹配：子任务号 + SKU编码
    cache_key = (task_no_str, sku_str)
    if cache_key in _sku_pick_time_cache:
        return _sku_pick_time_cache[cache_key], True

    return None, False


def _short_team(name):
    """班组简称：去前缀去后缀，保留核心区分词"""
    if not name:
        return ''
    s = name
    # 提取(B套)标记
    suffix = ''
    if '(B套)' in s:
        suffix = 'B'
        s = s.replace('(B套)', '')
    # 去掉常见前缀
    if s.startswith("自营标品"):
        s = s[4:]
    elif s.startswith("输送线"):
        # 输送线系列：去掉前缀，保留"中组""乙组"，不加B套后缀
        s = s[3:]
        return s or name
    # 去掉末尾的"组"
    if s.endswith('组'):
        s = s[:-1]
    # "蔬菜投线" 简化为 "蔬菜"
    if s == "蔬菜投线":
        s = "蔬菜"
    return (s + suffix) or name


def _get_sku_price(session, sku_code, warehouse_id="428"):
    """通过报损单/报溢单接口获取SKU售价"""
    try:
        now = datetime.now()
        start = (now - timedelta(days=365)).strftime("%Y-%m-%d 00:00:00")
        end = now.strftime("%Y-%m-%d 23:59:59")

        # 方式1：从报损单查价格
        res = request_with_retry(session, "/haina/billcenter/scrappedBill/r/pageList", {
            "skuCode": str(sku_code),
            "wareHouseId": warehouse_id,
            "warehouseId": warehouse_id,
            "createTimeStart": start,
            "createTimeEnd": end,
            "pageNo": 1,
            "pageSize": 1
        })
        if res and res.get("code") == 200:
            bills = res.get("data", {}).get("pageContent", [])
            if bills:
                bill_no = bills[0].get("scrappedBillNo")
                if bill_no:
                    detail_res = request_with_retry(session, "/haina/billcenter/scrappedBill/r/detail", {
                        "scrappedBillNo": bill_no,
                        "wareHouseId": warehouse_id,
                        "warehouseId": warehouse_id
                    })
                    if detail_res and detail_res.get("code") == 200:
                        details = detail_res.get("data", {}).get("details", [])
                        for d in details:
                            if str(d.get("skuCode")) == str(sku_code):
                                price = d.get("skuPrice")
                                if price:
                                    try:
                                        return round(float(price), 2)
                                    except (ValueError, TypeError):
                                        pass

        # 方式2：从报溢单查价格
        res2 = request_with_retry(session, "/haina/billcenter/overflowBill/r/pageList", {
            "skuCode": str(sku_code),
            "wareHouseId": warehouse_id,
            "warehouseId": warehouse_id,
            "createTimeStart": start,
            "createTimeEnd": end,
            "pageNo": 1,
            "pageSize": 1
        })
        if res2 and res2.get("code") == 200:
            bills2 = res2.get("data", {}).get("pageContent", [])
            if bills2:
                bill_no2 = bills2[0].get("overflowBillNo")
                if bill_no2:
                    detail_res2 = request_with_retry(session, "/haina/billcenter/overflowBill/r/detail", {
                        "overflowBillNo": bill_no2,
                        "wareHouseId": warehouse_id,
                        "warehouseId": warehouse_id
                    })
                    if detail_res2 and detail_res2.get("code") == 200:
                        details2 = detail_res2.get("data", {}).get("details", [])
                        for d in details2:
                            if str(d.get("skuCode")) == str(sku_code):
                                price = d.get("skuPrice")
                                if price:
                                    try:
                                        return round(float(price), 2)
                                    except (ValueError, TypeError):
                                        pass
        return None
    except Exception:
        return None


@bp.route('/blame_query')
def blame_query_page():
    return render_template('blame_query.html')


@bp.route('/api/blame_query', methods=['GET'])
def blame_query():
    """错标责任人查询：五步安全联动下钻，支持订单号或包裹号查询"""
    order_no = request.args.get('orderNo', '').strip()
    wh_id = request.args.get('whId', '').strip()
    parcel_no_input = request.args.get('parcelNo', '').strip()

    # 判断是否为包裹号查询模式
    is_parcel_mode = bool(parcel_no_input)

    if is_parcel_mode:
        # 包裹号模式：从包裹号解析仓ID
        # 格式：250010026000000360103362，前4位为前缀，接下来3位为仓ID
        if len(parcel_no_input) < 7:
            return jsonify({"ok": False, "msg": "包裹号格式不正确，长度不足"})
        parsed_wh_id = parcel_no_input[4:7]  # 提取仓ID（第5-7位）
        # 去掉前导零，如 "100" -> "100"，但 "001" -> "1"
        parsed_wh_id = str(int(parsed_wh_id))
        wh_id = wh_id or parsed_wh_id  # 优先使用用户手动选择的仓，否则用解析的
        parcel_no = parcel_no_input
    else:
        # 订单号模式：必须提供订单号和调拨仓
        if not order_no or not wh_id:
            return jsonify({"ok": False, "msg": "缺少订单号或目标调拨仓ID"})
        parcel_no = None

    token = get_access_token()
    if not token:
        return jsonify({"ok": False, "msg": "SSO Token 不可用，请确认本机 MOA 已登录"})

    session = get_shared_session()

    # 智能解析日期：订单号模式从订单号推算，包裹号模式使用当前日期范围
    if not is_parcel_mode:
        match = re.search(r'[A-Za-z]+(\d{6})', order_no)
        if match:
            try:
                order_date = datetime.strptime(match.group(1), "%y%m%d")
            except ValueError:
                order_date = datetime.now()
        else:
            order_date = datetime.now()
    else:
        # 包裹号模式：尝试从包裹号前缀推算日期（25xx -> 2025年），否则用当前日期
        order_date = datetime.now()

    appointment_date = (order_date + timedelta(days=1)).strftime("%Y-%m-%d")
    create_time_start = f"{order_date.strftime('%Y-%m-%d')} 00:00:00"
    create_time_end = f"{appointment_date} 23:59:59"
    stock_date = order_date.strftime("%Y-%m-%d")

    try:
        base_params = {
            "warehouseId": "428",
            "wareHouseId": "428",
        }

        # 1. 查包裹（包裹号模式跳过此步）
        sku_name = None
        sku_code = None
        sku_unit_desc = None
        brand_name = ""

        if not is_parcel_mode:
            params1_base = {**base_params, "orderNo": order_no, "allotInWarehouseId": wh_id,
                       "isFulfillmentWaveGray": "false", "fulfillmentWaveId": "", "pageNo": 1, "pageSize": 20}

            # 先用推算日期查询
            params1 = {**params1_base, "appointmentTime": appointment_date}
            res1 = request_with_retry(session, "/haina/ojs/rdc/r/parcelList", params1)

            page_content1 = []
            if res1 and isinstance(res1.get("data"), dict):
                page_content1 = res1["data"].get("pageContent", [])

            # 如果推算日期没查到，先不传日期试一次
            if not page_content1:
                res1 = request_with_retry(session, "/haina/ojs/rdc/r/parcelList", params1_base)
                if res1 and isinstance(res1.get("data"), dict):
                    page_content1 = res1["data"].get("pageContent", [])

            # 仍然没查到，按前后7天遍历
            if not page_content1:
                for delta in range(1, 8):
                    for sign in [-1, 1]:
                        try_date = (order_date + timedelta(days=delta * sign)).strftime("%Y-%m-%d")
                        params1_try = {**params1_base, "appointmentTime": try_date}
                        res1 = request_with_retry(session, "/haina/ojs/rdc/r/parcelList", params1_try)
                        if res1 and isinstance(res1.get("data"), dict):
                            page_content1 = res1["data"].get("pageContent", [])
                            if page_content1:
                                break
                    if page_content1:
                        break

            if not page_content1:
                return jsonify({"ok": False, "msg": f"未查询到包裹 (推算履约日:{appointment_date}，已尝试前后7天)，请核对仓位或单号"})

            target_parcel = page_content1[0]
            parcel_no = target_parcel.get("parcelNo")
            sku_name = target_parcel.get("skuName")
            sku_code = target_parcel.get("skuCode")
            sku_unit_desc = target_parcel.get("skuUnitDesc")
            brand_name = target_parcel.get("skuBrand") or target_parcel.get("brandName") or ""

            if not parcel_no:
                return jsonify({"ok": False, "msg": "包裹号提取失败，单据结构可能异常"})

        # 2. 查任务号
        # 辅助函数：判断接口返回是否有效且包含数据
        def _has_parcel_data(r):
            if not r:
                return False
            d = r.get("data")
            if isinstance(d, dict) and d.get("pageContent"):
                return True
            return False

        params2_base = {**base_params, "parcelNo": parcel_no, "allotInWarehouseId": wh_id,
                        "isFulfillmentWaveGray": "false", "fulfillmentWaveId": "",
                        "pageNo": 1, "pageSize": 20}

        res2 = None
        if is_parcel_mode:
            # 包裹号模式：先不传appointmentTime尝试，再按日期遍历
            params2_no_date = {**params2_base}
            res2 = request_with_retry(session, "/haina/ojs/rdc/r/containerParcelPrintList", params2_no_date)

            if not _has_parcel_data(res2):
                # 尝试前后7天的日期
                for delta in range(0, 8):
                    for sign in [0, -1, 1] if delta == 0 else [-1, 1]:
                        if delta == 0 and sign == 0:
                            continue
                        try_date = (datetime.now() + timedelta(days=delta * sign if sign != 0 else 0)).strftime("%Y-%m-%d")
                        params2_with_date = {**params2_base, "appointmentTime": try_date}
                        res2 = request_with_retry(session, "/haina/ojs/rdc/r/containerParcelPrintList", params2_with_date)
                        if _has_parcel_data(res2):
                            break
                    if _has_parcel_data(res2):
                        break

            if not _has_parcel_data(res2):
                debug_msg = f"包裹 {parcel_no} 查询失败(解析目的仓ID:{wh_id})"
                if res2:
                    debug_msg += f"，接口返回code:{res2.get('code')}, data类型:{type(res2.get('data')).__name__}"
                else:
                    debug_msg += "，接口无响应(Token可能失效)"
                return jsonify({"ok": False, "msg": debug_msg})
        else:
            # 订单号模式：先传appointmentTime，失败则重试
            params2 = {**params2_base, "appointmentTime": appointment_date}
            res2 = request_with_retry(session, "/haina/ojs/rdc/r/containerParcelPrintList", params2)

            if not _has_parcel_data(res2):
                # 先不传日期试一次
                res2 = request_with_retry(session, "/haina/ojs/rdc/r/containerParcelPrintList", params2_base)

            if not _has_parcel_data(res2):
                # 按前后7天遍历
                for delta in range(1, 8):
                    for sign in [-1, 1]:
                        try_date = (order_date + timedelta(days=delta * sign)).strftime("%Y-%m-%d")
                        params2_try = {**params2_base, "appointmentTime": try_date}
                        res2 = request_with_retry(session, "/haina/ojs/rdc/r/containerParcelPrintList", params2_try)
                        if _has_parcel_data(res2):
                            break
                    if _has_parcel_data(res2):
                        break

            if not _has_parcel_data(res2):
                return jsonify({"ok": False, "msg": f"包裹 {parcel_no} 任务交互失败（已尝试多日期查询）"})

        page_content2 = res2.get("data", {}).get("pageContent", [])
        if not page_content2:
            return jsonify({"ok": False, "msg": f"包裹({parcel_no})尚未生成拣货任务"})

        target_task = page_content2[0]
        zone_pick_bill_no = target_task.get("zonePickBillNo")
        picker_name = target_task.get("pickerName")
        receptacle_code = target_task.get("receptacleCode") or ""  # 容器码，用于匹配复核人

        # 包裹号模式：从第2步返回数据中补全SKU信息和订单号
        if is_parcel_mode:
            sku_name = sku_name or target_task.get("skuName")
            sku_code = sku_code or target_task.get("skuCode")
            sku_unit_desc = sku_unit_desc or target_task.get("skuUnitDesc")
            brand_name = brand_name or target_task.get("skuBrand") or target_task.get("brandName") or ""
            order_no = order_no or target_task.get("orderNo") or ""

            # 如果仍未获取到订单号，尝试通过parcelList接口用包裹号反查
            if not order_no:
                try:
                    params_reverse_base = {**base_params, "parcelNo": parcel_no, "allotInWarehouseId": wh_id,
                                           "isFulfillmentWaveGray": "false", "fulfillmentWaveId": "",
                                           "pageNo": 1, "pageSize": 5}

                    # 先不传appointmentTime尝试
                    res_reverse = request_with_retry(session, "/haina/ojs/rdc/r/parcelList", params_reverse_base)
                    reverse_content = []
                    if res_reverse and isinstance(res_reverse.get("data"), dict):
                        reverse_content = res_reverse["data"].get("pageContent", [])

                    # 如果没数据，按日期遍历（前后3天）
                    if not reverse_content:
                        for delta in [0, -1, 1, -2, 2, -3, 3]:
                            try_appt = (datetime.now() + timedelta(days=delta)).strftime("%Y-%m-%d")
                            params_with_date = {**params_reverse_base, "appointmentTime": try_appt}
                            res_reverse = request_with_retry(session, "/haina/ojs/rdc/r/parcelList", params_with_date)
                            if res_reverse and isinstance(res_reverse.get("data"), dict):
                                reverse_content = res_reverse["data"].get("pageContent", [])
                                if reverse_content:
                                    break

                    if reverse_content:
                        order_no = reverse_content[0].get("orderNo") or ""
                        if not sku_name:
                            sku_name = reverse_content[0].get("skuName")
                        if not sku_code:
                            sku_code = reverse_content[0].get("skuCode")
                        if not sku_unit_desc:
                            sku_unit_desc = reverse_content[0].get("skuUnitDesc")
                        if not brand_name:
                            brand_name = reverse_content[0].get("skuBrand") or reverse_content[0].get("brandName") or ""
                except Exception:
                    pass  # 反查失败不影响主流程

        if not zone_pick_bill_no:
            return jsonify({"ok": False, "msg": "未能下钻到该包裹的任务单号(zonePickBillNo)"})

        # 3. 查拣货单（包裹号模式扩大日期范围以确保能查到）
        if is_parcel_mode:
            search_start = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d 00:00:00")
            search_end = (datetime.now() + timedelta(days=1)).strftime("%Y-%m-%d 23:59:59")
        else:
            search_start = create_time_start
            search_end = create_time_end
        params3 = {**base_params, "zonePickBillNo": zone_pick_bill_no, "createTimeStart": search_start,
                   "createTimeEnd": search_end, "pageNo": 1, "pageSize": 20}
        res3 = request_with_retry(session, "/haina/outbound/zonepick/r/pageList", params3)

        handler_name = picker_name
        pick_time_str = None
        pick_time_source = None  # 记录时间来源：'sku_cache' 或 'task_complete'
        sku_pick_time_missing = False  # 标记是否缺少SKU级别时间

        if res3 and isinstance(res3.get("data"), dict):
            page_content3 = res3["data"].get("pageContent", [])
            if page_content3:
                handler_name = page_content3[0].get("handlerName") or picker_name
                # 从拣货子任务中提取容器码（用于后续查复核人）
                receptacle_code = page_content3[0].get("receptacleCode") or receptacle_code

                # 优先从导出文件缓存中获取SKU级别的商品拣货时间
                if zone_pick_bill_no and sku_code:
                    cached_time, found = get_sku_pick_time(zone_pick_bill_no, sku_code)
                    if found:
                        pick_time_str = cached_time
                        pick_time_source = "sku_cache"
                        if is_parcel_mode and pick_time_str:
                            try:
                                pick_dt = datetime.strptime(pick_time_str, "%Y-%m-%d %H:%M:%S")
                                stock_date = pick_dt.strftime("%Y-%m-%d")
                            except Exception:
                                pass

                # 如果缓存未命中，回退到子任务级别的 completeTime
                if not pick_time_str:
                    sku_pick_time_missing = True
                    finish_ts = page_content3[0].get("completeTime") or page_content3[0].get("lastModifyTime")
                    if finish_ts:
                        try:
                            pick_dt = datetime.fromtimestamp(int(finish_ts) / 1000)
                            pick_time_str = pick_dt.strftime("%Y-%m-%d %H:%M:%S")
                            pick_time_source = "task_complete"
                            if is_parcel_mode:
                                stock_date = pick_dt.strftime("%Y-%m-%d")
                        except Exception:
                            pick_time_str = None

        # 辅助函数：查询人员班组/类型信息（warehouse_id 指定查哪个仓的人员名单）
        def _get_person_type_info(person_name, warehouse_id="428"):
            if not person_name:
                return ""
            try:
                user_params = {
                    "name": person_name,
                    "warehouseValidity": "EFFECTIVE",
                    "warehouseIdList": str(warehouse_id),
                    "jobStatus": "INCUMBENCY",
                    "pageNo": 1,
                    "pageSize": 20,
                }
                res_user = request_with_retry(session, "/hrm/labour/inhouse/user/r/pageUserList", user_params)
                if res_user and isinstance(res_user.get("data"), dict):
                    user_list = res_user["data"].get("pageContent", [])
                    user_data = None
                    for u in user_list:
                        if u.get("name") == person_name:
                            user_data = u
                            break
                    if user_data:
                        service_desc = str(user_data.get("labourServiceTypeDesc", ""))
                        team_name = user_data.get("teamOrgName", "")
                        if "临时" in service_desc:
                            return "[临]"
                        elif team_name:
                            return f"[{_short_team(team_name)}]"
            except Exception:
                pass
            return ""

        # 查询拣货人员类型（固定工/临时工）及班组信息
        picker_type_info = _get_person_type_info(handler_name)

        # 4. 查子明细
        params4 = {**base_params, "pickBillNo": zone_pick_bill_no, "pageNo": 1, "pageSize": 100}
        res4 = request_with_retry(session, "/haina/outbound/zonepick/r/detail", params4)

        if not res4 or not isinstance(res4.get("data"), dict):
            return jsonify({"ok": False, "msg": f"拣货单明细 {zone_pick_bill_no} 提取失败"})

        details = res4["data"].get("pageContent", [])

        # 包裹号模式：如果前面未获取到SKU信息，从明细中取第一条作为目标商品
        if is_parcel_mode and not sku_code and details:
            first_detail = details[0]
            sku_code = first_detail.get("skuCode")
            sku_name = sku_name or first_detail.get("skuName")
            sku_unit_desc = sku_unit_desc or first_detail.get("skuUnitDesc")
            brand_name = brand_name or first_detail.get("skuBrand") or ""

        # 5. 查当天库位库存快照
        location_code = None
        if sku_code:
            for d in details:
                if str(d.get("skuCode")) == str(sku_code):
                    location_code = d.get("locationCode")
                    break

        stock_details = []
        if location_code:
            params5 = {
                **base_params,
                "locationCode": location_code,
                "canPick": "ALL",
                "stockType": "NORMAL",
                "freezeQtyEnum": "ALL",
                "date": stock_date,
                "pageNo": 1,
                "pageSize": 20
            }
            res5 = request_with_retry(session, "/haina/stock/aggregation/stocksummarydate/r/pageList", params5)
            if res5 and isinstance(res5.get("data"), dict):
                stock_details = res5["data"].get("pageContent", [])

        # 构建提醒信息
        pick_time_note = ""
        if sku_pick_time_missing:
            order_date_str = order_date.strftime("%m月%d日")
            pick_time_note = f"⚠️ 未找到相关记录，请手动下载{order_date_str}的储位拣货子任务详情"

        # 6. 通过容器码查询调拨复核人员
        checker_name = ""
        checker_type_info = ""
        if receptacle_code:
            try:
                check_time_start = (order_date - timedelta(days=1)).strftime("%Y-%m-%d") + " 00:00:00"
                check_time_end = (order_date + timedelta(days=3)).strftime("%Y-%m-%d") + " 23:59:59"
                check_params = {
                    **base_params,
                    "containerCode": receptacle_code,
                    "createTimeStart": check_time_start,
                    "createTimeEnd": check_time_end,
                    "pageNo": 1,
                    "pageSize": 20
                }
                res_check = request_with_retry(session, "/haina/ocs/allot/r/allotCheck/list", check_params)
                if res_check and isinstance(res_check.get("data"), dict):
                    check_list = res_check["data"].get("pageContent", [])
                    # 找到容器码匹配的记录，取 checkTaskNo 查详情日志获取复核人
                    check_task_no = None
                    for chk in check_list:
                        if chk.get("containerCode") == receptacle_code and chk.get("checkTaskNo"):
                            check_task_no = chk["checkTaskNo"]
                            break
                    if check_task_no:
                        detail_params = {
                            **base_params,
                            "checkTaskNo": check_task_no,
                        }
                        res_detail = request_with_retry(session, "/haina/ocs/allot/r/allotCheck/detail", detail_params)
                        if res_detail and isinstance(res_detail.get("data"), dict):
                            log_list = res_detail["data"].get("allotCheckTaskLogList", [])
                            # 优先取"完成复核任务"的操作人，其次取"领取复核任务"的操作人
                            receive_name = ""
                            complete_name = ""
                            for log in log_list:
                                operation = log.get("operation", "")
                                if operation == "完成复核任务" and log.get("createByName"):
                                    complete_name = log["createByName"]
                                elif operation == "领取复核任务" and log.get("createByName"):
                                    receive_name = log["createByName"]
                            checker_name = complete_name or receive_name
            except Exception:
                pass  # 复核人查询失败不影响主流程

        # 查询复核人员班组信息（复核人在凤岗仓428工作）
        if checker_name:
            checker_type_info = _get_person_type_info(checker_name)

        return jsonify({
            "ok": True,
            "data": {
                "order_info": {
                    "order_no": order_no,
                    "parcel_no": parcel_no,
                    "sku_code": sku_code,
                    "sku_name": sku_name,
                    "brand_name": brand_name,
                    "sku_unit_desc": sku_unit_desc,
                    "picker_name": handler_name,
                    "picker_type_info": picker_type_info,
                    "zone_pick_bill_no": zone_pick_bill_no,
                    "pick_time": pick_time_str,
                    "pick_time_source": pick_time_source,
                    "pick_time_note": pick_time_note,
                    "target_location": location_code,
                    "stock_date": stock_date,
                    "checker_name": checker_name,
                    "checker_type_info": checker_type_info,
                    "receptacle_code": receptacle_code
                },
                "details": details,
                "stock_details": stock_details
            }
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"ok": False, "msg": f"后端追查发生异常错误: {str(e)}"})


@bp.route('/api/ticket_query', methods=['GET'])
def ticket_query():
    """根据订单号查询客诉工单信息"""
    order_no = request.args.get('orderNo', '').strip()
    if not order_no:
        return jsonify({"ok": False, "msg": "缺少订单号"})

    token = get_access_token()
    if not token:
        return jsonify({"ok": False, "msg": "SSO Token 不可用"})

    session = get_shared_session()

    match = re.search(r'[A-Za-z]+(\d{6})', order_no)
    if match:
        try:
            order_date = datetime.strptime(match.group(1), "%y%m%d")
        except ValueError:
            order_date = datetime.now()
    else:
        order_date = datetime.now()

    date_from = (order_date - timedelta(days=2)).strftime("%Y-%m-%d")
    date_to = (order_date + timedelta(days=7)).strftime("%Y-%m-%d")

    try:
        params = {
            "responsibleParty": "",
            "csuName": "",
            "orderNo": order_no,
            "ticketId": "",
            "ticketCreateDateFrom": date_from,
            "ticketCreateDateTo": date_to,
            "pageNo": 1,
            "pageSize": 20,
        }
        res = request_with_retry(session, "/tms/ticket/bill/r/pagelist", params)

        if not res or not isinstance(res.get("data"), dict):
            return jsonify({"ok": True, "data": {"tickets": []}})

        tickets = res["data"].get("pageContent", [])

        result_tickets = []
        for t in tickets:
            create_ts = t.get("ticketCreateTime")
            create_time_str = None
            if create_ts:
                try:
                    create_time_str = datetime.fromtimestamp(int(create_ts) / 1000).strftime("%Y-%m-%d %H:%M:%S")
                except Exception:
                    pass

            result_tickets.append({
                "ticket_id": t.get("ticketId"),
                "create_time": create_time_str,
                "warehouse_name": t.get("warehouseName"),
                "business_name": t.get("businessName"),
                "driver_name": t.get("driverName"),
                "sku_name": t.get("csuName"),
                "sku_unit_desc": t.get("csuUnitDesc"),
                "problem_name": t.get("problemName"),
                "reason_name": t.get("reasonName"),
                "question_desc": t.get("questionDesc"),
                "process_result": t.get("processResult"),
                "responsible_party": t.get("responsibleParty"),
                "detail_url": f"https://klwms.meituan.com/app/operation/responsibility/ticketManagement/detail?ticketId={t.get('ticketId')}"
            })

        return jsonify({"ok": True, "data": {"tickets": result_tickets}})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"ok": False, "msg": f"工单查询异常: {str(e)}"})


@bp.route('/api/ticket_detail', methods=['GET'])
def ticket_detail():
    """根据工单ID查询工单详情（含图片）"""
    ticket_id = request.args.get('ticketId', '').strip()
    if not ticket_id:
        return jsonify({"ok": False, "msg": "缺少工单ID"})

    token = get_access_token()
    if not token:
        return jsonify({"ok": False, "msg": "SSO Token 不可用"})

    session = get_shared_session()

    try:
        params = {"ticketId": ticket_id}
        res = request_with_retry(session, "/tms/ticket/bill/r/ticketDetail", params)

        if not res or res.get("code") != 200 or not isinstance(res.get("data"), dict):
            return jsonify({"ok": False, "msg": "工单详情获取失败"})

        data = res["data"]
        return jsonify({
            "ok": True,
            "data": {
                "ticket_id": data.get("ticketId"),
                "problem_name": data.get("problemName"),
                "question_desc": data.get("questionDesc"),
                "process_result": data.get("processResult"),
                "pic_urls": data.get("picUrls", []),
                "ticket_status": data.get("ticketStatus"),
                "business_name": data.get("businessName"),
                "driver_name": data.get("driverName"),
                "sku_list": data.get("ticketSkuList", []),
            }
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"ok": False, "msg": f"工单详情查询异常: {str(e)}"})


@bp.route('/api/sku_location_search', methods=['GET'])
def sku_location_search():
    """按商品名/品牌模糊搜索可拣货位"""
    keyword = request.args.get('skuName', '').strip()
    brand_keyword = request.args.get('skuBrand', '').strip()
    order_sku_code = request.args.get('orderSkuCode', '').strip()
    stock_date = request.args.get('date', '').strip()
    warehouse_id = "428"

    if not keyword and not brand_keyword:
        return jsonify({"ok": False, "msg": "请输入商品名称或品牌关键字"})

    token = get_access_token()
    if not token:
        return jsonify({"ok": False, "msg": "SSO Token 不可用"})

    session = get_shared_session()

    try:
        base_params = {
            "warehouseId": warehouse_id,
            "wareHouseId": warehouse_id,
            "canPick": "YES",
            "stockType": "NORMAL",
            "freezeQtyEnum": "ALL",
            "pageNo": 1,
            "pageSize": 50
        }
        if stock_date:
            base_params["date"] = stock_date

        search_params = {**base_params}
        if keyword:
            search_params["skuName"] = keyword
        if brand_keyword:
            search_params["skuBrand"] = brand_keyword

        res = request_with_retry(session, "/haina/stock/aggregation/stocksummarydate/r/pageList", search_params)

        if not res or not isinstance(res.get("data"), dict):
            return jsonify({"ok": False, "msg": "库存查询接口异常"})

        items = res["data"].get("pageContent", [])

        results = []
        for item in items:
            available = item.get("availableQty", 0) or 0
            if available > 0:
                if brand_keyword:
                    item_brand = item.get("skuBrand") or ""
                    item_name = item.get("skuName") or ""
                    if brand_keyword not in item_brand and brand_keyword not in item_name:
                        continue
                results.append({
                    "location_code": item.get("locationCode"),
                    "sku_code": item.get("skuCode"),
                    "sku_name": item.get("skuName"),
                    "sku_brand": item.get("skuBrand"),
                    "sku_unit_desc": item.get("skuUnitDesc"),
                    "available_qty": available,
                    "area_name": item.get("areaName"),
                    "category_name": item.get("categoryName"),
                })

        # 查询订单商品的可拣货位信息
        order_sku_info = None
        if order_sku_code:
            order_params = {**base_params, "skuCode": order_sku_code}
            res2 = request_with_retry(session, "/haina/stock/aggregation/stocksummarydate/r/pageList", order_params)
            if res2 and isinstance(res2.get("data"), dict):
                order_items = res2["data"].get("pageContent", [])
                valid_order_items = [i for i in order_items if (i.get("availableQty") or 0) > 0]
                if valid_order_items:
                    best = max(valid_order_items, key=lambda x: x.get("availableQty", 0))
                    order_sku_info = {
                        "location_code": best.get("locationCode"),
                        "sku_code": best.get("skuCode"),
                        "sku_name": best.get("skuName"),
                        "sku_brand": best.get("skuBrand"),
                        "sku_unit_desc": best.get("skuUnitDesc"),
                        "category_name": best.get("categoryName"),
                        "area_name": best.get("areaName"),
                    }

        # 查询商品售价
        order_price = None
        if order_sku_code:
            order_price = _get_sku_price(session, order_sku_code, warehouse_id)
            if order_sku_info:
                order_sku_info["sku_price"] = order_price

        queried_prices = {}
        sku_codes_to_query = list(set(str(r["sku_code"]) for r in results))
        for code in sku_codes_to_query[:10]:
            price = _get_sku_price(session, code, warehouse_id)
            if price is not None:
                queried_prices[code] = price

        for r in results:
            r["sku_price"] = queried_prices.get(str(r["sku_code"]))

        return jsonify({"ok": True, "data": {
            "results": results,
            "total": len(results),
            "order_sku_info": order_sku_info
        }})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"ok": False, "msg": f"库位查询异常: {str(e)}"})
